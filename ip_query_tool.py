import requests
import time
import openpyxl
from openpyxl.styles import Font, Border, Side
import urllib3
import os
import ipaddress

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Fetch the API key and context from environment variables
api_key = os.getenv('API_KEY')
context = os.getenv('CONTEXT')

# Check if API_KEY and CONTEXT are set
if not api_key or not context:
    print("Error: API_KEY or CONTEXT is not set in the environment.")
    exit(1)

# URL and headers
url = 'https://api.radwarecloud.app/api/v1/sdcc/threat/core/insight/_bulkResolve'
headers = {
    'x-api-key': api_key,
    'Context': context,
    'Content-Type': 'application/json'
}

# Batch size
batch_size = 70

# Retry settings
max_retries = 5
initial_retry_delay = 5  # seconds

# Function to prompt the user to create the file if not found
def prompt_user_to_create_file():
    print("Input file 'ip_list.txt' not found.")
    print("Please create a file named 'ip_list.txt' in the script directory.")
    print("Add a list of IP addresses, one per line, and save the file.")
    input("Press Enter after creating the file to continue...")

# Try to read the list of IP addresses from the file, prompt user if not found
while True:
    try:
        with open('ip_list.txt', 'r') as file:
            ip_addresses = [line.strip() for line in file.readlines()]
        break  # Exit the loop if the file was read successfully
    except FileNotFoundError:
        prompt_user_to_create_file()

# Function to validate IP addresses
def is_valid_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

# Filter valid IP addresses
valid_ip_addresses = [ip for ip in ip_addresses if is_valid_ip(ip)]
invalid_ip_addresses = [ip for ip in ip_addresses if not is_valid_ip(ip)]

if invalid_ip_addresses:
    print(f"Invalid IP addresses found and will be excluded: {', '.join(invalid_ip_addresses)}")

# Split IP addresses into batches
ip_batches = [valid_ip_addresses[i:i + batch_size] for i in range(0, len(valid_ip_addresses), batch_size)]

# Dictionary to hold all responses
all_responses = {}

# Set to hold all unique headers
all_headers = set()

# Loop through each batch of IP addresses
for i, batch in enumerate(ip_batches):
    print(f"Processing batch {i + 1}/{len(ip_batches)} with {len(batch)} IPs...")

    # Prepare the payload with all IPs in the current batch
    payload = {
        "addresses": batch,
        "projection": ["all"]
    }

    attempt = 0
    retry_delay = initial_retry_delay
    while attempt < max_retries:
        try:
            # Send one request for the entire batch
            response = requests.post(url, headers=headers, json=payload, verify=False, timeout=30)

            # Check if the request was successful
            if response.status_code == 200:
                response_data = response.json()

                # Process the data for each IP in the batch
                for ip in batch:
                    # Check if 'results' and the IP are in the response data
                    if 'results' in response_data and ip in response_data['results']:
                        all_responses[ip] = response_data['results'][ip]
                    else:
                        print(f"Data for IP {ip} not found in the response.")
                        all_responses[ip] = {}  # Assign an empty dict for missing data

                    # Update the set of all headers with the keys of the current IP's data
                    all_headers.update(all_responses[ip].keys())
                break  # Exit retry loop if successful

            else:
                print(f"Failed to get data for batch {i + 1}. Status code: {response.status_code}")
                if response.status_code == 400:
                    print(f"Response Content: {response.content.decode()}")
                attempt += 1
                time.sleep(retry_delay)  # Wait before retrying
                retry_delay *= 2  # Exponential backoff

        except requests.exceptions.RequestException as e:
            print(f"Error while requesting data for batch {i + 1}: {e}")
            if response is not None:
                print(f"Response Status Code: {response.status_code}")
                print(f"Response Content: {response.content.decode()}")
            attempt += 1
            time.sleep(retry_delay)  # Wait before retrying
            retry_delay *= 2  # Exponential backoff

    # Optional: Pause between batches to avoid server overload
    time.sleep(1)  # Pause for 1 second between batches

# Convert set of headers to a sorted list and replace underscores with spaces
all_headers = ["IP"] + sorted([header.replace('_', ' ') for header in all_headers])

# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "IP Data"

# Define styles for headers and borders
header_font = Font(bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Capitalize headers and write them to the first row with formatting
for col, header in enumerate(all_headers, 1):
    cell = ws.cell(row=1, column=col, value=header.capitalize())
    cell.font = header_font
    cell.border = thin_border

# Write data for each IP with borders
for row, (ip, data) in enumerate(all_responses.items(), start=2):
    data['IP'] = ip  # Include the IP address in the row
    for col, header in enumerate(all_headers, 1):
        # Convert header back to original form (with underscores) to retrieve data
        original_header = header.replace(' ', '_')
        value = data.get(original_header, '')
        # Handle list values by converting them to strings
        if isinstance(value, list):
            value = ', '.join(value)
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border

# Save the Excel file before modifications
excel_file = 'ip_data.xlsx'
wb.save(excel_file)

# Re-open the workbook to modify columns
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Identify the column index of "Risk score" (with space)
risk_score_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Risk score":
        risk_score_col = col
        break

# If "Risk score" column is found, move it to the second position
if risk_score_col:
    risk_score_values = []
    risk_score_styles = []
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=risk_score_col)
        risk_score_values.append(cell.value)
        risk_score_styles.append(cell._style)  # Capture the style of each cell

    # Delete the original "Risk score" column
    ws.delete_cols(risk_score_col)

    # Insert the "Risk score" values into the second column with their original styles
    ws.insert_cols(2)
    for row, (value, style) in enumerate(zip(risk_score_values, risk_score_styles), start=1):
        cell = ws.cell(row=row, column=2, value=value)
        cell._style = style  # Apply the original style to the new cell

# Save the modified Excel file
wb.save(excel_file)
print(f"Data has been modified and written to {excel_file}")
